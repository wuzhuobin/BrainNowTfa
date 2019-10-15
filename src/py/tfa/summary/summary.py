#%%
from os import path
import pandas as pd
from matplotlib import pyplot
import openpyxl as xl

input_xlsxs = ['test data output.xlsx']
output = 'result.xlsx'

#%%
def summary(input_xlsxs, output):
  pass
#%%
  # work sheets needed to be generated.
  group_summary = pd.DataFrame()
  vlf_lf_hf = pd.DataFrame()
  work_sheets = {
    'l_gain': pd.DataFrame(),
    'l_phase': pd.DataFrame(),
    'l_coherence': pd.DataFrame(),
    'r_gain': pd.DataFrame(),
    'r_phase': pd.DataFrame(),
    'r_coherence': pd.DataFrame(),
  }
  # column_header are generated from input filenames.
  filename_prefix = list(map(lambda fileName: path.splitext(path.basename(fileName))[0], input_xlsxs))
  for i in range(len(work_sheets)):
    pass
    key = list(work_sheets.keys())[i]
    for j in range(len(input_xlsxs)):
      pass
      data_frame = pd.read_excel(input_xlsxs[j])
      if j == 0:
        work_sheets[key]['F'] = data_frame['F']
      if i == 0 & j == 0:
        group_summary['F'] = data_frame['F']
      work_sheets[key][filename_prefix[j]] = data_frame[key]
    # calculate the mean value
    work_sheets[key]['Average'] = work_sheets[key][filename_prefix].mean(axis=1)
    group_summary[key] = work_sheets[key]['Average']
    
  with pd.ExcelWriter(output) as writer:
    pass
    group_summary.to_excel(writer, sheet_name='Group summary', index=False)
    # write an empty work sheet for place holding
    vlf_lf_hf.to_excel(writer, sheet_name='VLF_LF_HF', index=False)
    for key in work_sheets:
      pass
      work_sheets[key].to_excel(writer, sheet_name=key, index=False)
#%% using openpyxl to manipulate data
  wb = xl.load_workbook(filename=output)
  ws_vlf_lf_hf = wb['VLF_LF_HF']
  ws_vlf_lf_hf['A3'] = 'VLF (0.02-0.07 Hz)'
  ws_vlf_lf_hf['A4'] = 'Gain, %/mmHg'
  ws_vlf_lf_hf['A5'] = 'Phase, radian'
  ws_vlf_lf_hf['A6'] = 'Coherence'
  ws_vlf_lf_hf['A7'] = 'LF (0.07-0.20 Hz)'
  ws_vlf_lf_hf['A8'] = 'Gain, %/mmHg'
  ws_vlf_lf_hf['A9'] = 'Phase, radian'
  ws_vlf_lf_hf['A10'] = 'Coherence'
  ws_vlf_lf_hf['A11'] = 'HF (0.20-0.35 Hz)'
  ws_vlf_lf_hf['A12'] = 'Gain, %/mmHg'
  ws_vlf_lf_hf['A13'] = 'Phase, radian'
  ws_vlf_lf_hf['A14'] = 'Coherence'

  for i in range(len(filename_prefix)):
    pass
    ws_vlf_lf_hf.cell(row=1, column=i * 2 + 2, value=filename_prefix[i])
    ws_vlf_lf_hf.cell(row=2, column=i * 2 + 2, value='L side')
    ws_vlf_lf_hf.cell(row=2, column=i * 2 + 3, value='R side')

  for i in range(len(work_sheets)):
    pass
    key = list(work_sheets.keys())[i]
    for j in range(len(filename_prefix)):
      pass
      vlf = work_sheets[key].query('F > 0.02 & F < 0.07').mean()
      lf = work_sheets[key].query('F > 0.07 & F < 0.20').mean()
      hf = work_sheets[key].query('F > 0.20 & F < 0.35').mean()
      vlf_lf_hf[key + '_vlf'] = vlf
      vlf_lf_hf[key + '_lf'] = lf
      vlf_lf_hf[key + '_hf'] = hf
      if i < 3:
        pass
        ws_vlf_lf_hf.cell(row=i + 4, column=j * 2 + 2, value=vlf[filename_prefix[j]])
        ws_vlf_lf_hf.cell(row=i + 8, column=j * 2 + 2, value=lf[filename_prefix[j]])
        ws_vlf_lf_hf.cell(row=i + 12, column=j * 2 + 2, value=hf[filename_prefix[j]])
      else:
        pass
        ws_vlf_lf_hf.cell(row=i + 1, column=j * 2 + 3, value=vlf[filename_prefix[j]])
        ws_vlf_lf_hf.cell(row=i + 5, column=j * 2 + 3, value=lf[filename_prefix[j]])
        ws_vlf_lf_hf.cell(row=i + 9, column=j * 2 + 3, value=hf[filename_prefix[j]])
      # Adding summary to input_xlsxs.
      wb_input_xlsxs = xl.load_workbook(filename=input_xlsxs[j])
      ws_input_xlsxs = wb_input_xlsxs.active
      ws_input_xlsxs['K1'] = 'L side'
      ws_input_xlsxs['L1'] = 'R side'
      ws_input_xlsxs['J2'] = 'VLF (0.02-0.07 Hz)'
      ws_input_xlsxs['J3'] = 'Gain, %/mmHg'
      ws_input_xlsxs['J4'] = 'Phase, radian'
      ws_input_xlsxs['J5'] = 'Coherence'
      ws_input_xlsxs['J6'] = 'LF (0.07-0.20 Hz)'
      ws_input_xlsxs['J7'] = 'Gain, %/mmHg'
      ws_input_xlsxs['J8'] = 'Phase, radian'
      ws_input_xlsxs['J9'] = 'Coherence'
      ws_input_xlsxs['J10'] = 'HF (0.20-0.35 Hz)'
      ws_input_xlsxs['J11'] = 'Gain, %/mmHg'
      ws_input_xlsxs['J12'] = 'Phase, radian'
      ws_input_xlsxs['J13'] = 'Coherence'
      if i < 3:
        pass
        ws_input_xlsxs.cell(row=i + 3, column=11, value=vlf[filename_prefix[j]])
        ws_input_xlsxs.cell(row=i + 7, column=11, value=lf[filename_prefix[j]])
        ws_input_xlsxs.cell(row=i + 11, column=11, value=hf[filename_prefix[j]])
      else:
        ws_input_xlsxs.cell(row=i + 0, column=12, value=vlf[filename_prefix[j]])
        ws_input_xlsxs.cell(row=i + 4, column=12, value=lf[filename_prefix[j]])
        ws_input_xlsxs.cell(row=i + 8, column=12, value=hf[filename_prefix[j]])
        pass
      wb_input_xlsxs.save(input_xlsxs[j])


  
  ws_vlf_lf_hf.cell(row=1, column=len(filename_prefix) * 2 + 2, value='Average')
  ws_vlf_lf_hf.cell(row=2, column=len(filename_prefix) * 2 + 2, value='L side')
  ws_vlf_lf_hf.cell(row=2, column=len(filename_prefix) * 2 + 3, value='R side')
  vlf_lf_hf_average = vlf_lf_hf.drop(['F', 'Average']).mean()

  for i in range(len(work_sheets)):
    pass
    key = list(work_sheets.keys())[i]
    if i < 3:
      pass
      ws_vlf_lf_hf.cell(row=i + 4, column=len(filename_prefix) * 2 + 2, value=vlf_lf_hf_average[key + '_vlf'])
      ws_vlf_lf_hf.cell(row=i + 8, column=len(filename_prefix) * 2 + 2, value=vlf_lf_hf_average[key + '_lf'])
      ws_vlf_lf_hf.cell(row=i + 12, column=len(filename_prefix) * 2 + 2, value=vlf_lf_hf_average[key + '_hf'])
    else:
      pass
      ws_vlf_lf_hf.cell(row=i + 1, column=len(filename_prefix) * 2 + 3, value=vlf_lf_hf_average[key + '_vlf'])
      ws_vlf_lf_hf.cell(row=i + 5, column=len(filename_prefix) * 2 + 3, value=vlf_lf_hf_average[key + '_lf'])
      ws_vlf_lf_hf.cell(row=i + 9, column=len(filename_prefix) * 2 + 3, value=vlf_lf_hf_average[key + '_hf'])

  ws_group_summary = wb['Group summary']
  ws_group_summary['J1'] = 'L side'
  ws_group_summary['K1'] = 'R side'
  ws_group_summary['I2'] = 'VLF (0.02-0.07 Hz)'
  ws_group_summary['I3'] = 'Gain, %/mmHg'
  ws_group_summary['I4'] = 'Phase, radian'
  ws_group_summary['I5'] = 'Coherence'
  ws_group_summary['I6'] = 'LF (0.07-0.20 Hz)'
  ws_group_summary['I7'] = 'Gain, %/mmHg'
  ws_group_summary['I8'] = 'Phase, radian'
  ws_group_summary['I9'] = 'Coherence'
  ws_group_summary['I10'] = 'HF (0.20-0.35 Hz)'
  ws_group_summary['I11'] = 'Gain, %/mmHg'
  ws_group_summary['I12'] = 'Phase, radian'
  ws_group_summary['I13'] = 'Coherence'

  ws_group_summary['J3'] = vlf_lf_hf_average['l_gain_vlf']
  ws_group_summary['K3'] = vlf_lf_hf_average['r_gain_vlf']
  ws_group_summary['J4'] = vlf_lf_hf_average['l_phase_vlf']
  ws_group_summary['K4'] = vlf_lf_hf_average['r_phase_vlf']
  ws_group_summary['J5'] = vlf_lf_hf_average['l_coherence_vlf']
  ws_group_summary['K5'] = vlf_lf_hf_average['r_coherence_vlf']

  ws_group_summary['J7'] = vlf_lf_hf_average['l_gain_lf']
  ws_group_summary['K7'] = vlf_lf_hf_average['r_gain_lf']
  ws_group_summary['J8'] = vlf_lf_hf_average['l_phase_lf']
  ws_group_summary['K8'] = vlf_lf_hf_average['r_phase_lf']
  ws_group_summary['J9'] = vlf_lf_hf_average['l_coherence_lf']
  ws_group_summary['K9'] = vlf_lf_hf_average['r_coherence_lf']

  ws_group_summary['J11'] = vlf_lf_hf_average['l_gain_hf']
  ws_group_summary['K11'] = vlf_lf_hf_average['r_gain_hf']
  ws_group_summary['J12'] = vlf_lf_hf_average['l_phase_hf']
  ws_group_summary['K12'] = vlf_lf_hf_average['r_phase_hf']
  ws_group_summary['J13'] = vlf_lf_hf_average['l_coherence_hf']
  ws_group_summary['K13'] = vlf_lf_hf_average['r_coherence_hf']

  wb.save(filename=output)

#%%
  fig = pyplot.figure(figsize=[8.27, 11.69])
  pyplot.subplots_adjust(wspace=0.2, hspace=0.3)
  pyplot.subplot(3, 2, 1)
  pyplot.plot(group_summary.F, group_summary.l_gain)
  pyplot.title('l_gain')
  pyplot.subplot(3, 2, 3)
  pyplot.plot(group_summary.F, group_summary.l_phase)
  pyplot.title('l_phase')
  pyplot.subplot(3, 2, 5)
  pyplot.plot(group_summary.F, group_summary.l_coherence)
  pyplot.title('l_coherence')
  pyplot.subplot(3, 2, 2)
  pyplot.plot(group_summary.F, group_summary.r_gain)
  pyplot.title('r_gain')
  pyplot.subplot(3, 2, 4)
  pyplot.plot(group_summary.F, group_summary.r_phase)
  pyplot.title('r_phase')
  pyplot.subplot(3, 2, 6)
  pyplot.plot(group_summary.F, group_summary.r_coherence)
  pyplot.title('r_coherence')
  fig.savefig(output + '.png')
